from openpyxl import load_workbook
import os
import gen_SDC as gen_SDC
import apps as apps
import recognizer
path = os.getcwd()
grammar = path + "\\VoA\\" +"Grammar_en_GB.xlsx"



    
    def processUtterance(self,text):
        result_Intent="Nothing Understood"
        prompt="I am afraid I dont know what you mean"
        recog = ""
        wb = load_workbook(grammar)
        sheet_intent = wb.get_sheet_by_name("Intent")
        print type(sheet_intent),sheet_intent.max_row,type(sheet_intent.max_row)
        for rowidx in range(sheet_intent.max_row):
            if sheet_intent.cell(rowidx,2).value == text:
                print sheet_intent.cell(rowidx,2).value
                print "Utt= "+ text + " Intent= " + sheet_intent.cell(rowidx,1).value
        
                result_Intent = sheet_intent.cell(rowidx,1).value
                print result_Intent,rowidx
                
                if sheet_intent.cell(rowidx,4) is not None:
                    print sheet_intent.cell(rowidx,4).value
                    # try:
                    recog = eval(sheet_intent.cell(rowidx,4).value)
                break
    ###########################
        sheet_prompt = wb.get_sheet_by_name("Prompt")
        print "*********"
        if result_Intent!="Nothing Understood":
            for row in range(sheet_prompt.max_row):
                if sheet_prompt.cell(row,1).value == result_Intent:
                    print sheet_prompt.cell(row,2).value
                    print "Intent= "+ result_Intent + " Prompt= " + sheet_prompt.cell(row,2).value + recog
                    prompt = sheet_prompt.cell(row,2).value + recog
                    return prompt
                    break
        else:
            prompt = "I did not understand that"
