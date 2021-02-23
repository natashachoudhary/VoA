import speech_recognition as sr 
import pyttsx3
from openpyxl import load_workbook
import os,time
import gen_SDC as gen_SDC
import apps as apps
import recognizer
from fuzzywuzzy import fuzz
path = os.getcwd()
grammar = path + "\\VoA\\" +"Grammar_en_GB.xlsx"

class recognizer():
    def __init__(self):
        self.mic_name="Microphone (Realtek(R) Audio)"
        self.sample_rate = 48000
        self.chunk_size=2048
        self.r = sr.Recognizer() 

    def eInit(self):
        mic_list = sr.Microphone.list_microphone_names() 
        for i, microphone_name in enumerate(mic_list): 
            if microphone_name == self.mic_name: 
                self.device_id = i
    
    def eListening(self):
        with sr.Microphone(device_index = self.device_id, sample_rate = self.sample_rate, chunk_size = self.chunk_size) as source: 
            #wait for a second to let the recognizer adjust the  
            #energy threshold based on the surrounding noise level 
            #self.r.adjust_for_ambient_noise(source) 
            self.ePrompting("Say Something " )
            #listens for the user's input 
            self.audio = self.r.listen(source)
            return self.audio

    def eProcessing(self):
        recognitionResult = 0
        try:  
            self.text = self.r.recognize_google(self.audio)
            print("you said " + self.text )
            recognitionResult=1
        #error occurs when google could not understand what was said 
        except sr.UnknownValueError: 
            self.text ="Sorry I did not hear it. Please repeat"
            self.ePrompting(self.text)
            time.sleep(0.5)
        except sr.RequestError as e: 
                print ("Could not request results from Google Speech Recognition service; {0}".format(e)) 
        return self.text,recognitionResult

    def ePrompting(self,prompt):
        try:
            # Initialize the engine 
            engine = pyttsx3.init() 
            # engine.setProperty('voice', voices[1].id)
            # engine.setProperty('volume',1.0)
            engine.say(prompt)  
            engine.runAndWait() 
            return 0
        except:
            return 1
    
    def processUtterance(self,text):
        result_Intent="Nothing Understood"
        prompt="I am afraid I dont know what you mean"
        recog = ""
        wb = load_workbook(grammar)
        sheet_intent = wb.get_sheet_by_name("Intent")
        #print type(sheet_intent),sheet_intent.max_row,type(sheet_intent.max_row)
        for rowidx in range(1,sheet_intent.max_row+1):
            if fuzz.token_set_ratio(self.text,sheet_intent.cell(rowidx,2).value)>=80:
                print sheet_intent.cell(rowidx,2).value
                print "Utt= "+ self.text + " Intent= " + sheet_intent.cell(rowidx,1).value
        
                result_Intent = sheet_intent.cell(rowidx,1).value
                #print result_Intent,rowidx
                
                if sheet_intent.cell(rowidx,4).value==None:
                    recog = ""
                else:
                    print sheet_intent.cell(rowidx,4).value, type(sheet_intent.cell(rowidx,4).value)
                    try:
                        recog = eval(sheet_intent.cell(rowidx,4).value)
                        recog = str(recog)
                        break
                    except :
                        recog = "Error occured"
                        print "Error"
        ###########################
        sheet_prompt = wb.get_sheet_by_name("Prompt")
        print "*********"
        if result_Intent!="Nothing Understood":
            for row in range(1,sheet_prompt.max_row+1):
                if sheet_prompt.cell(row,1).value == result_Intent:
                    #print sheet_prompt.cell(row,2).value
                    print "Intent= "+ result_Intent + " Prompt= " + sheet_prompt.cell(row,2).value + recog
                    prompt = sheet_prompt.cell(row,2).value + recog
                    return prompt
                    break
        else:
            prompt = "I did not understand that"